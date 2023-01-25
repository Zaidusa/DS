import openpyxl
import os

filename = "../Testdata/Testdata.xlsx"
sheetname = "std"
class excelread:

    @staticmethod
    def rowcount(filename,sheetname):
        wb=openpyxl.load_workbook(filename)
        sheet=wb[sheetname]
        return(sheet.max_row)

    @staticmethod
    def columncount(filename, sheetname):
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheetname]
        return(sheet.max_column)

    @staticmethod
    def readcelldata(filename, sheetname,rownum,colnum):
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheetname]
        value=sheet.cell(row=rownum, column=colnum).value
        print (value)

    @staticmethod
    def readdata(filename, sheetname):

        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheetname]
        dict = {'username': [], 'password': []}
        for i in range(2,sheet.max_row):

            for j in range(1,sheet.max_column):
                #print(i,j)
                value = str(sheet.cell(row=i, column=j).value)
                #print(i,j)
                value1 = str(sheet.cell(row=i, column=j+1).value)
                dict['username'].append(value)
                dict['password'].append(value1)
        return (dict)

excelread.readdata(filename,sheetname)

